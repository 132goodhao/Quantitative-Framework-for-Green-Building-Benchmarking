# -*- coding: utf-8 -*-
"""
Created on Thu Sep 14 17:57:01 2023

@author: goodhao
"""

import pandas as pd
import glob
import os
import openpyxl

variables = ['Building Name', 'Address', 'Type', 'Size', 'TOP/CSC year', 'Green Mark Award', \
             'Year of GMA', 'GM Version', 'GFA', 'AC Area', 'AC Area Percentage', 'Type of ACS', \
                 'Age of Chiller', 'ACS efficiency', 'Date of Last Audit/Health Check', 'LED Percentage Usage', 'PV Installation']
    
input_path = r'01_Original' 
output_folder = r'02_Reformed_Original'
file_path = r'01_Original\ttt\Listing of Building Energy Performance Data for 2021.xlsx' 

files1 = glob.glob(input_path + '\Listing of Building Energy Performance Data for 2021.xlsx')
wb = openpyxl.load_workbook(files1[0])
sheet = wb.active


current_variables = [cell.value for cell in sheet[1]]
# 遍历给定的变量列表
for i, variable in enumerate(variables):
    if variable in current_variables:
        # 找到变量名称在当前文件中的索引，并将列移动到正确的位置
        index = current_variables.index(variable) + 1
        sheet.move_range(f"A{i+1}:A{i+1}", cols=index-1)

        # 生成新的变量名称，并在第一行的相应位置写入
        new_variable_name = variable
        for j in range(index-1):
            new_variable_name = " " + new_variable_name

        sheet.cell(row=1, column=index, value=new_variable_name)

        # 从当前变量列表中移除已经重命名的变量名称
        current_variables.remove(variable)

# 将剩余变量按照首字母升序排在后面
current_variables.sort()
for variable in current_variables:
    column = len(sheet[1]) + 1
    sheet.cell(row=1, column=column, value=variable)
    
new_file_name = os.path.splitext(file_path)[0] + " VarRenamed.xlsx"
wb.save(new_file_name)
wb.close() 
