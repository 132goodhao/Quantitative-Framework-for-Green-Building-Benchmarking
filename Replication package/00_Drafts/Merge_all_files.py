# -*- coding: utf-8 -*-
"""
Created on Thu Sep 14 20:50:56 2023

@author: goodhao
"""

import os
import pandas as pd
from openpyxl import load_workbook

# 批量读取文件夹里面的excel文件
folder_path = r'03_Reformed_Reformed'
output_file = r'04_Reformed_and_Merged\Energy Performance Data from 2016 to 2021.xlsx'

# 获取文件夹里所有xlsx文件的路径
file_paths = [os.path.join(folder_path, file) for file in os.listdir(folder_path) if file.endswith('.xlsx')]

# 读取第一个文件作为目标文件
target_file_path = file_paths[0]
target_wb = load_workbook(target_file_path)
target_ws = target_wb.active

# 将其他文件的数据合并到目标文件后面
for file_path in file_paths[1:]:
    wb = load_workbook(file_path)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):  # 从第二行开始迭代
        values = [cell.value for cell in row]
        target_ws.append(values)
        
# 保存合并后的文件
target_wb.save(output_file)
